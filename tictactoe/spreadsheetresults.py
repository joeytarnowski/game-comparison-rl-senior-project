from openpyxl import Workbook
from openpyxl.styles import Font
import pickle
"""
Helper script to convert all results into a spreadsheet
"""
def fill_results(sheet, agents, name, names, results):
    bold_font = Font(bold=True)
    italics_font = Font(italic=True)
    # Total Results
    sheet["A1"] = f"{name} Results"
    sheet["A2"] = "Wins"
    sheet["A3"] = "Losses"
    sheet["A4"] = "Draws"
    for x in range(4):
        sheet[f"A{x+1}"].font = bold_font

    i = 0
    for agent in agents:
        calc_total = float(sum(results[i][0]) + sum(results[i][1]) + sum(results[i][2]))
        sheet.cell(row=1, column=i+2).value = names[i]
        sheet.cell(row=1, column=i+2).font = italics_font
        for x in range(3):
            percent_calc = 100*float(sum(results[i][x]))/calc_total
            sheet.cell(row=x+2, column=i+2).value = f"{sum(results[i][x])} ({round(percent_calc, ndigits=2)}%)"
        i += 1

    for i in range(max(len(agent.testing_results_rand[0]), len(agent2.testing_results_rand[0]), len(agent3.testing_results_rand[0]), len(agent4.testing_results_rand[0]))):
        # Create a new section for each test
        sheet.cell(row=6+(i*5), column=1).value = f"Test {i+1} Results"
        sheet.cell(row=7+(i*5), column=1).value = "Wins"
        sheet.cell(row=8+(i*5), column=1).value = "Losses"
        sheet.cell(row=9+(i*5), column=1).value = "Draws"
        for x in range(4):
            sheet.cell(row=6+x+(i*5), column=1).font = bold_font

        j = 0
        win_total = 0
        loss_total = 0
        draw_total = 0
        for agent in agents:
            calc_total = float(results[j][0][i] + results[j][1][i] + results[j][2][i])
            # Print agent results
            sheet.cell(row=6+(i*5), column=j+2).value = names[j]
            sheet.cell(row=6+(i*5), column=j+2).font = italics_font
            for x in range(3):
                percent_calc = 100*float(results[j][x][i])/calc_total
                sheet.cell(row=7+x+(i*5), column=j+2).value = f"{results[j][0][i]} ({round(percent_calc, ndigits=2)}%)"
            # Add each result to total results
            win_total += results[j][0][i]
            loss_total += results[j][1][i]
            draw_total += results[j][2][i]
            j += 1
        calc_total = float(win_total + loss_total + draw_total)
        # Print total results
        totals = [win_total, loss_total, draw_total]
        sheet.cell(row=6+(i*5), column=6).value = "Total Results"
        sheet.cell(row=6+(i*5), column=6).font = italics_font
        for x in range(3):
            percent_calc = 100*float(totals[x])/calc_total
            sheet.cell(row=7+x+(i*5), column=6).value = f"{totals[x]} ({round(percent_calc, ndigits=2)}%)"
    
        
workbook = Workbook()
total_sheet = workbook.active
total_sheet.title = "Total Results"
rand_sheet = workbook.create_sheet(title="Random Results")
opt_sheet = workbook.create_sheet(title="Optimal Results")

with open('q_agent.pkl', 'rb') as f:
    agent = pickle.load(f)
with open('sarsa_agent.pkl', 'rb') as g:
    agent2 = pickle.load(g)
with open('mcoff_agent.pkl', 'rb') as h:
    agent3 = pickle.load(h)
with open('mcon_agent.pkl', 'rb') as i:
    agent4 = pickle.load(i)

agents = [agent, agent2, agent3, agent4]
names = ["Q-Learning", "Sarsa", "MC Off-Policy", "MC On-Policy"]
total_results = [agent.testing_results_rand+agent.testing_results_opt, agent2.testing_results_rand+agent2.testing_results_opt, agent3.testing_results_rand+agent3.testing_results_opt, agent4.testing_results_rand+agent4.testing_results_opt]
rand_results = [agent.testing_results_rand, agent2.testing_results_rand, agent3.testing_results_rand, agent4.testing_results_rand]
opt_results = [agent.testing_results_opt, agent2.testing_results_opt, agent3.testing_results_opt, agent4.testing_results_opt]

fill_results(total_sheet, agents, "Total", names, total_results)
fill_results(rand_sheet, agents, "Random", names, rand_results)
fill_results(opt_sheet, agents, "Optimal", names, opt_results)

workbook.save(filename="tictactoe.xlsx")